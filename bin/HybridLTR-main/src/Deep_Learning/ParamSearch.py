import argparse
import os
import sys

current_folder = os.path.dirname(os.path.abspath(__file__))
# Add the path to the 'configs' folder to the Python path
configs_folder = os.path.join(current_folder, "../..")
sys.path.append(configs_folder)


import random
from configs import config, gpu_config
from DataProcessor import DataProcessor
from src.Deep_Learning.Classifier import Classifier
from src.Deep_Learning.Trainer import Trainer
from src.Deep_Learning.CrossValidator import CrossValidator
from utils.data_util import get_feature_len, get_gpu_config
from utils.show_util import showTrainParams

def test_use_all_repbase_no_tsd():
    # 测试分别使用不同的特征变化，五折交叉验证
    features = [(1, 1, 0, 1, 1), (1, 1, 1, 1, 0), (1, 0, 1, 1, 1), (0, 1, 1, 1, 1), (1, 1, 1, 0, 1)]
    for feature in features:
        #加载数据前，先设定使用哪些特征，并定义好特征维度

        print('feature:', feature)
        config.use_terminal = feature[0]
        config.use_TSD = feature[1]
        config.use_domain = feature[2]
        config.use_ends = feature[3]
        config.use_kmers = feature[4]


        X_feature_len = 0
        if config.use_kmers != 0:
            for kmer_size in config.internal_kmer_sizes:
                X_feature_len += pow(4, kmer_size)
            if config.use_terminal != 0:
                for i in range(2):
                    for kmer_size in config.terminal_kmer_sizes:
                        X_feature_len += pow(4, kmer_size)
        if config.use_TSD != 0:
            X_feature_len += config.max_tsd_length * 4 + 1
        if config.use_domain != 0:
            X_feature_len += 29
        if config.use_ends != 0:
            X_feature_len += 10 * 4
        config.X_feature_len = X_feature_len

        data_processor = DataProcessor()
        # 加载数据
        # 请确保下面数据的header格式为Repbase格式，即'TE_name  Superfamily Species'，以'\t'分割
        cv_train_data_path = config.work_dir + "/repbase_total.ref"  # 交叉验证训练数据路径
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes, cv_train_data_path)
        print(X.shape, y.shape)

        validator = CrossValidator(num_folds=5)

        means, stdvs = validator.evaluate(X, y)
        print('accuracy, precision, recall, f1:')
        print("Mean array:", means)
        print("stdv array:", stdvs)

def test_use_features():
    # 测试分别使用不同的特征变化，五折交叉验证
    features = [(0, 0, 0, 0), (0, 1, 0, 0), (1, 1, 0, 0), (1, 1, 0, 1), (1, 1, 1, 1)]
    #加载数据前，先设定使用哪些特征，并定义好特征维度
    for feature in features:
        print('feature:', feature)
        config.use_terminal = feature[0]
        config.use_TSD = feature[1]
        config.use_domain = feature[2]
        config.use_ends = feature[3]


        X_feature_len = 0
        for kmer_size in config.internal_kmer_sizes:
            X_feature_len += pow(4, kmer_size)
        if config.use_terminal != 0:
            for i in range(2):
                for kmer_size in config.terminal_kmer_sizes:
                    X_feature_len += pow(4, kmer_size)
        if config.use_TSD != 0:
            X_feature_len += config.max_tsd_length * 4 + 1
        if config.use_domain != 0:
            X_feature_len += 29
        if config.use_ends != 0:
            X_feature_len += 10 * 4
        config.X_feature_len = X_feature_len


        data_processor = DataProcessor()
        # 加载数据
        # 请确保下面数据的header格式为Repbase格式，即'TE_name  Superfamily Species'，以'\t'分割
        cv_train_data_path = config.work_dir + "/repbase_total.ref"  # 交叉验证训练数据路径
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes, cv_train_data_path)
        print(X.shape, y.shape)

        validator = CrossValidator(num_folds=5)

        means, stdvs = validator.evaluate(X, y)
        print('accuracy, precision, recall, f1:')
        print("Mean array:", means)
        print("stdv array:", stdvs)


def test_cnn_kernel_size(work_dir):
    # 固定kmer_size, 卷积层数, kernels数量, 测试单个kernel里的kernel size的变化对结果的影响
    # 创建一个新的工作簿
    import openpyxl
    workbook = openpyxl.Workbook()

    config.work_dir = work_dir
    config.is_predict = 0
    config.use_minority = 0

    # 读取 Excel 文件
    file_path = config.work_dir + '/cnn_kernel_size.xlsx'
    worksheet = workbook.active
    data = [
        ['CNN kernel size', 'Accuracy', 'Precision', 'Recall', 'F1']
    ]
    # 逐行写入数据，并在每次循环后保存
    for row in data:
        worksheet.append(row)
        workbook.save(file_path)

    # 打开已存在的工作簿
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    cnn_kernel_size_arrays = [3, 5, 7, 9, 11]
    for i, cnn_kernel_size in enumerate(cnn_kernel_size_arrays):
        config.cnn_kernel_sizes_array = 3 * [cnn_kernel_size]
        print('cnn_kernel_sizes_array:', config.cnn_kernel_sizes_array)

        X_feature_len = get_feature_len()
        config.X_feature_len = X_feature_len

        data_path = config.work_dir + "/train.ref"
        print('data_path:', data_path)
        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes,
                                                              data_path)
        print(X.shape, y.shape)

        trainer = Trainer()

        model_path = trainer.train(X, y, config.cnn_num_convs, config.cnn_filters_array)
        # print('Trained model is stored in:', model_path)

        data_path = config.work_dir + "/valid.ref"
        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes,
                                                              data_path)
        print(X.shape, y.shape)

        classifier = Classifier(model_path=model_path)

        accuracy, precision, recall, f1 = classifier.predict(X, y, seq_names, data_path)
        print('accuracy, precision, recall, f1:', accuracy, precision, recall, f1)
        # 逐行写入数据，并在每次循环后保存
        row = [str(cnn_kernel_size), str(accuracy), str(precision), str(recall), str(f1)]
        worksheet.append(row)
        workbook.save(file_path)

def test_cnn_filters(work_dir):
    # 固定kmer_size, 卷积层数, 测试kernels数量的变化对结果的影响
    # 创建一个新的工作簿
    import openpyxl
    workbook = openpyxl.Workbook()

    config.work_dir = work_dir
    config.is_predict = 0
    config.use_minority = 0

    file_path = config.work_dir + '/cnn_kernels.xlsx'
    worksheet = workbook.active
    data = [
        ['CNN kernels', 'Accuracy', 'Precision', 'Recall', 'F1']
    ]
    for row in data:
        worksheet.append(row)
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    cnn_filters_arrays = [4, 8, 16, 32, 64, 128]
    for i, cnn_filters_array in enumerate(cnn_filters_arrays):
        config.cnn_filters_array = 3 * [cnn_filters_array]
        print('cnn_filters_array:', config.cnn_filters_array)

        params = {}
        params['data_path'] = work_dir
        params['outdir'] = work_dir
        params['genome'] = work_dir
        showTrainParams(params)

        X_feature_len = get_feature_len()
        config.X_feature_len = X_feature_len

        data_path = config.work_dir + "/train.ref"
        print('data_path:', data_path)
        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes,
                                                              data_path)
        print(X.shape, y.shape)

        trainer = Trainer()

        model_path = trainer.train(X, y, config.cnn_num_convs, config.cnn_filters_array)
        # print('Trained model is stored in:', model_path)

        data_path = config.work_dir + "/valid.ref"
        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes,
                                                              data_path)
        print(X.shape, y.shape)

        classifier = Classifier(model_path=model_path)

        accuracy, precision, recall, f1 = classifier.predict(X, y, seq_names, data_path)
        print('accuracy, precision, recall, f1:', accuracy, precision, recall, f1)
        row = [str(cnn_filters_array), str(accuracy), str(precision), str(recall), str(f1)]
        worksheet.append(row)
        workbook.save(file_path)

def test_cnn_layers(work_dir):
    # 固定kmer_size, 滤波器数量， 测试卷积层数的变化对结果的影响
    # 创建一个新的工作簿
    import openpyxl
    workbook = openpyxl.Workbook()

    config.work_dir = work_dir
    config.is_predict = 0
    config.use_minority = 0

    file_path = config.work_dir + '/cnn_num.xlsx'
    worksheet = workbook.active
    data = [
        ['CNN number', 'Accuracy', 'Precision', 'Recall', 'F1']
    ]
    for row in data:
        worksheet.append(row)
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    cnn_num_convs_array = [1, 2, 3, 4, 5, 6, 7, 8]
    for i, cnn_num_convs in enumerate(cnn_num_convs_array):
        config.cnn_num_convs = cnn_num_convs
        config.cnn_filters_array = cnn_num_convs * [32]
        config.cnn_kernel_sizes_array = cnn_num_convs * [3]
        print('cnn_num_convs:', config.cnn_num_convs)

        X_feature_len = get_feature_len()
        config.X_feature_len = X_feature_len

        data_path = config.work_dir + "/train.ref"
        print('data_path:', data_path)
        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes,
                                                              data_path)
        print(X.shape, y.shape)

        trainer = Trainer()


        model_path = trainer.train(X, y, config.cnn_num_convs, config.cnn_filters_array)
        # print('Trained model is stored in:', model_path)


        data_path = config.work_dir + "/valid.ref"
        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes,
                                                              data_path)
        print(X.shape, y.shape)

        classifier = Classifier(model_path=model_path)

        accuracy, precision, recall, f1 = classifier.predict(X, y, seq_names, data_path)
        print('accuracy, precision, recall, f1:', accuracy, precision, recall, f1)
        row = [str(cnn_num_convs), str(accuracy), str(precision), str(recall), str(f1)]
        worksheet.append(row)
        workbook.save(file_path)

def test_kmer_size():
    # 即固定卷积层数和滤波器数量，测试kmer组合的变化对结果的影响，此时只使用kmer特征
    kmer_sizes_array = [
                        [1, 2], [1, 3], [1, 4], [1, 5],
                        [2, 3], [2, 4], [2, 5], [3, 4], [3, 5],
                        [1, 2, 3], [1, 2, 4], [1, 2, 5], [1, 3, 4], [1, 3, 5], [2, 3, 4], [2, 3, 5], [3, 4, 5],
                        [1, 2, 3, 4], [1, 2, 3, 5], [1, 2, 4, 5], [1, 3, 4, 5], [2, 3, 4, 5]]
    for kmer_sizes in kmer_sizes_array:
        print('kmer sizes:', kmer_sizes)
        X_feature_len = 0
        for kmer_size in kmer_sizes:
            X_feature_len += pow(4, kmer_size)
        if config.use_terminal != 0:
            for i in range(2):
                for kmer_size in kmer_sizes:
                    X_feature_len += pow(4, kmer_size)
        if config.use_TSD != 0:
            X_feature_len += 11 * 4 + 1
        if config.use_domain != 0:
            X_feature_len += 29
        if config.use_ends != 0:
            X_feature_len += 10 * 4
        config.X_feature_len = X_feature_len

        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(kmer_sizes)
        print(X.shape, y.shape)

        trainer = Trainer()

        model_path = trainer.train(X, y, config.cnn_num_convs, config.cnn_filters_array)
        # print('Trained model is stored in:', model_path)

        data_processor = DataProcessor()
        X, y, seq_names, data_path, data_path = data_processor.load_data(kmer_sizes)
        print(X.shape, y.shape)

        model_path = config.project_dir + '/models/model.h5'
        classifier = Classifier(model_path=model_path)

        accuracy, precision, recall, f1 = classifier.predict(X, y)
        print('accuracy, precision, recall, f1:', accuracy, precision, recall, f1)

def test_kmer_size_combination(work_dir):
    import openpyxl
    workbook = openpyxl.Workbook()

    config.work_dir = work_dir
    config.is_predict = 0
    config.use_minority = 0

    file_path = config.work_dir + '/kmer_size_search/kmer_size_test.xlsx'
    # data = pd.read_excel(file_path)
    #
    # # 获取第一列数据
    # first_column = data.iloc[:, 0]  # 通过 iloc 方法选择所有行的第一列
    # second_column = data.iloc[:, 1]  # 通过 iloc 方法选择所有行的第二列
    # # 将每个字符串转换为列表
    # first_column = [ast.literal_eval(word) for word in first_column]
    # second_column = [ast.literal_eval(word) for word in second_column]
    combinations = []
    # for i in range(len(first_column)):
    #     combinations.append(first_column[i] + second_column[i])

    worksheet = workbook.active
    data = [
        ['Internal Kmer sizes', 'Terminal Kmer sizes', 'Accuracy', 'Precision', 'Recall', 'F1']
    ]

    for row in data:
        worksheet.append(row)
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)

    worksheet = workbook.active

    # 随机生成kmer_size的组合，限制在50种
    internal_kmer_sizes_array = [[2], [3], [4], [5], [1, 2], [1, 3], [1, 4], [1, 5], [2, 3], [2, 4], [2, 5], [3, 4],
                                 [3, 5],
                                 [1, 2, 3], [1, 2, 4], [1, 2, 5], [1, 3, 4], [1, 3, 5], [2, 3, 4], [2, 3, 5], [3, 4, 5]]
    terminal_kmer_sizes_array = [[2], [3], [4], [5], [1, 2], [1, 3], [1, 4], [1, 5], [2, 3], [2, 4], [2, 5], [3, 4],
                                 [3, 5],
                                 [1, 2, 3], [1, 2, 4], [1, 2, 5], [1, 3, 4], [1, 3, 5], [2, 3, 4], [2, 3, 5], [3, 4, 5]]
    max_num = 50


    i = 0
    while i < max_num:
        internal_kmer_sizes = list(random.choice(internal_kmer_sizes_array))
        terminal_kmer_sizes = list(random.choice(terminal_kmer_sizes_array))
        combination = internal_kmer_sizes + terminal_kmer_sizes
        if combination in combinations:
            continue
        else:
            combinations.append(combination)
            i += 1

        config.internal_kmer_sizes = internal_kmer_sizes
        config.terminal_kmer_sizes = terminal_kmer_sizes

        print('internal_kmer_sizes:', config.internal_kmer_sizes)
        print('terminal_kmer_sizes:', config.terminal_kmer_sizes)

        X_feature_len = get_feature_len()
        config.X_feature_len = X_feature_len

        data_path = config.work_dir + "/train.ref"
        print('data_path:', data_path)
        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes, data_path)
        print(X.shape, y.shape)

        trainer = Trainer()

        model_path = trainer.train(X, y, config.cnn_num_convs, config.cnn_filters_array)
        # print('Trained model is stored in:', model_path)

        data_path = config.work_dir + "/valid.ref"
        data_processor = DataProcessor()
        X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes, data_path)
        print(X.shape, y.shape)

        classifier = Classifier(model_path=model_path)

        accuracy, precision, recall, f1 = classifier.predict(X, y, seq_names, data_path)
        print('accuracy, precision, recall, f1:', accuracy, precision, recall, f1)
        # 逐行写入数据，并在每次循环后保存
        row = [str(internal_kmer_sizes), str(terminal_kmer_sizes), str(accuracy), str(precision), str(recall), str(f1)]
        worksheet.append(row)
        workbook.save(file_path)

#测试数据增强前后，各个类别的性能提升
def each_class_improve_use_data_augmentation():
    data_path = config.work_dir + "/repbase_total.no_rice.ref"
    data_processor = DataProcessor()
    X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes, data_path)
    print(X.shape, y.shape)

    trainer = Trainer()

    model_path = trainer.train(X, y, config.cnn_num_convs, config.cnn_filters_array)
    print('Trained model is stored in:', model_path)

    model_path = '/home/hukang/NeuralTE/models/model_2023-09-14.18-44-42.h5'

    data_path = config.work_dir + "/repbase_total.rice.ref"
    data_processor = DataProcessor()
    X, y, seq_names, data_path = data_processor.load_data(config.internal_kmer_sizes, config.terminal_kmer_sizes, data_path)
    print(X.shape, y.shape)


    classifier = Classifier(model_path=model_path)

    accuracy, precision, recall, f1 = classifier.predict(X, y, seq_names)
    print('accuracy, precision, recall, f1:', accuracy, precision, recall, f1)

def main():
    # 1.parse args
    describe_info = '########################## NeuralTE, version ' + str(config.version_num) + ' ##########################'
    parser = argparse.ArgumentParser(description=describe_info)
    parser.add_argument('--work_dir', required=True, metavar='work_dir',
                        help='Input current work dir.')


    args = parser.parse_args()

    work_dir = args.work_dir

    #prefix = '/public/home/hpc194701009'
    #work_dir = prefix + '/NeuralTE/data/param_tuning/Dataset1'

    gpu_config.start_gpu_num = 2
    gpu_config.use_gpu_num = 1

    # reload GPU config
    get_gpu_config(gpu_config.start_gpu_num, gpu_config.use_gpu_num)

    # file_dir = work_dir + '/kmer_size_search'
    # if not os.path.exists(file_dir):
    #     os.makedirs(file_dir)
    # test_kmer_size_combination(work_dir)

    # file_dir = work_dir + '/cnn_num_search'
    # if not os.path.exists(file_dir):
    #     os.makedirs(file_dir)
    # test_cnn_layers(work_dir)


    file_dir = work_dir + '/cnn_filters_search'
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    test_cnn_filters(work_dir)

    file_dir = work_dir + '/cnn_kernel_size_search'
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    test_cnn_kernel_size(work_dir)



if __name__ == '__main__':
    main()